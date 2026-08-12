# -*- coding: utf-8 -*-
"""Excel / CSV 导入导出与模板生成。"""
import csv
import io
import os
import re

from openpyxl import Workbook, load_workbook

HEADERS = ["公司名称", "联系人", "电话", "邮箱", "地区", "客户类型", "状态", "来源", "标签", "备注", "地址"]
HEADER_MAP = {
    "公司名称": "name", "公司": "name", "名称": "name", "name": "name",
    "联系人": "contact", "联系人姓名": "contact",
    "电话": "phone", "手机": "phone", "手机号": "phone", "联系电话": "phone",
    "邮箱": "email", "email": "email", "邮件": "email",
    "地区": "region", "区域": "region", "城市": "region",
    "客户类型": "type", "类型": "type",
    "状态": "status",
    "来源": "source", "线索来源": "source",
    "标签": "tags", "tag": "tags",
    "备注": "note", "备注说明": "note",
    "地址": "address", "公司地址": "address",
}

SOCIAL_HEADERS = ["平台", "作品/笔记标题", "作品链接", "评论人昵称", "评论内容", "评论时间", "备注"]
MAP_HEADERS = ["公司名称", "地址", "电话", "分类", "城市", "备注"]
MAP_HEADER_ALIASES = {
    "公司名称": "name", "Name": "name", "name": "name",
    "地址": "address", "Address": "address", "address": "address",
    "电话": "phone", "Phone": "phone", "PhoneNumber": "phone", "phone": "phone",
    "分类": "category", "Category": "category", "Type": "category", "category": "category",
    "城市": "city", "City": "city", "city": "city",
    "备注": "note", "Notes": "note", "note": "note",
    "网站": "website", "Website": "website", "website": "website",
    "评分": "rating", "Rating": "rating", "rating": "rating",
    "评论数": "reviews", "Reviews": "reviews", "reviews": "reviews",
}


def build_map_template_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "地图线索导入模板"
    ws.append(["公司名称", "地址", "电话", "分类", "城市", "网站", "评分", "评论数", "备注"])
    ws.append(["广州XX弱电工程有限公司", "广州市天河区xx路xx号", "020-88886666", "弱电工程", "广州", "https://example.com", "4.5", "120", "google-maps-scraper / 后羿采集器导出后整理"])
    for i, w in enumerate([32, 36, 16, 16, 12, 30, 8, 10, 46], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_map(path):
    """解析地图采集导出文件（.xlsx/.csv）→ 线索列表。"""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
    else:
        raw = open(path, "rb").read()
        text = None
        for enc in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            return [], "无法识别文件编码"
        rows = [r for r in csv.reader(io.StringIO(text))]
    if not rows:
        return [], "文件为空"
    header = [str(c or "").strip() for c in rows[0]]
    idx = {}
    for i, h in enumerate(header):
        if h in MAP_HEADER_ALIASES:
            idx.setdefault(MAP_HEADER_ALIASES[h], i)
    if "name" not in idx:
        return [], "表头缺少“公司名称/Name”（兼容 google-maps-scraper、后羿采集器导出）"

    def val(row, key):
        i = idx.get(key)
        if i is None or i >= len(row):
            return ""
        v = row[i].value if hasattr(row[i], "value") else row[i]
        return str(v or "").strip()

    leads = []
    for n, row in enumerate(rows[1:], start=2):
        name = val(row, "name")
        if not name:
            continue
        category = val(row, "category")
        from core.mapsearch import _map_type
        note = val(row, "note")
        extra = []
        if val(row, "website"):
            extra.append("网站：" + val(row, "website"))
        if val(row, "rating"):
            extra.append("评分：" + val(row, "rating"))
        if val(row, "reviews"):
            extra.append(val(row, "reviews") + " 条评价")
        if extra:
            note = (note + "；" if note else "") + "；".join(extra)
        leads.append({
            "name": val(row, "name"),
            "address": val(row, "address"),
            "phone": val(row, "phone"),
            "region": val(row, "city"),
            "type": _map_type(category) if category else "其他",
            "source": "地图获客",
            "tags": category or "地图",
            "note": note,
        })
    return leads, None


def build_social_template_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "社媒评论导入模板"
    ws.append(SOCIAL_HEADERS)
    ws.append(["抖音", "光缆熔接工艺讲解", "https://v.douyin.com/xxxx", "光纤小王", "这个熔接机多少钱？想了解", "2026-08-01 10:00", ""])
    ws.append(["小红书", "光纤收发器选购指南", "https://www.xiaohongshu.com/explore/xxxx", "弱电老张", "求报价，机房改造用", "2026-08-01 11:00", ""])
    for i, w in enumerate([10, 28, 34, 16, 42, 20, 20], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_wechat_template_txt():
    return (
        "# 微信聊天记录导入模板（.txt / .csv）\n"
        "# 每段记录格式：联系人 + 日期时间 + 换行 + 消息内容，支持以下两种格式：\n"
        "# 格式一（推荐，来自微信导出工具）：\n"
        "张三 2026-08-01 10:00:00\n"
        "你好，我们公司需要采购一批光缆，请问有报价吗？\n"
        "张三 2026-08-01 10:05:00\n"
        "另外光纤收发器也发一份价格表\n"
        "# 格式二：\n"
        "# 2026-08-01 10:00 李四: 消息内容\n"
    ).encode("utf-8-sig")


def parse_social(path):
    """解析社媒评论 CSV/XLSX → 线索列表。"""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
    else:
        raw = open(path, "rb").read()
        text = None
        for enc in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            return [], "无法识别文件编码"
        rows = [r for r in csv.reader(io.StringIO(text))]
    if not rows:
        return [], "文件为空"
    header = [str(c or "").strip() for c in rows[0]]
    idx = {h: i for i, h in enumerate(header) if h in SOCIAL_HEADERS}
    if "评论人昵称" not in idx:
        return [], "表头缺少“评论人昵称”（请使用：平台、作品/笔记标题、作品链接、评论人昵称、评论内容、评论时间）"

    def val(row, h):
        i = idx.get(h)
        if i is None or i >= len(row):
            return ""
        v = row[i].value if hasattr(row[i], "value") else row[i]
        return str(v or "").strip()

    leads = []
    for n, row in enumerate(rows[1:], start=2):
        nick = val(row, "评论人昵称")
        if not nick:
            continue
        platform = val(row, "平台")
        title = val(row, "作品/笔记标题")
        link = val(row, "作品链接")
        content = val(row, "评论内容")
        ctime = val(row, "评论时间")
        remark = val(row, "备注")
        note = f"评论：{content}"
        if title:
            note += f"\n作品：{title}"
        if link:
            note += f"\n链接：{link}"
        if remark:
            note += f"\n备注：{remark}"
        leads.append({
            "name": f"{nick}（{platform or '社媒'}）",
            "contact": nick,
            "region": "",
            "type": "终端客户",
            "status": "新线索",
            "source": "社媒评论",
            "tags": platform or "社媒",
            "note": note[:500],
            "reminder_date": ctime[:10] if ctime else "",
        })
    return leads, None


WX_HEADER_1 = re.compile(r"^\[?(\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}(?::\d{2})?)\]?\s*(.+?)[:：]\s*(.*)$")
WX_HEADER_2 = re.compile(r"^(.+?)\s+(\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}(?::\d{2})?)$")


def parse_wechat(path):
    """解析微信聊天记录文本 → 线索列表（按联系人分组）。"""
    raw = open(path, "rb").read()
    text = None
    for enc in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return [], "无法识别文件编码"
    chats = {}  # sender -> list of (time, content)
    current = None
    current_time = ""
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "聊天记录" in line and len(line) < 20:
            continue
        m1 = WX_HEADER_1.match(line)
        if m1:
            current = m1.group(2).strip()
            current_time = m1.group(1).replace("T", " ")
            content = m1.group(3).strip()
            if current:
                chats.setdefault(current, []).append((current_time, content or "（表情/图片）"))
            continue
        m2 = WX_HEADER_2.match(line)
        if m2:
            current = m2.group(1).strip()
            current_time = m2.group(2).replace("T", " ")
            continue
        if current:
            chats.setdefault(current, []).append((current_time, line))
    if not chats:
        return [], "没有解析到聊天记录。请使用模板格式：联系人+时间换行+消息内容，或“时间 联系人: 内容”"
    leads = []
    for sender, msgs in chats.items():
        if len(sender) > 40:
            continue
        lines = []
        for t, c in msgs[-30:]:
            lines.append((f"[{t}] " if t else "") + c)
        first_time = next((t for t, _ in msgs if t), "")
        leads.append({
            "name": sender,
            "contact": sender,
            "source": "微信记录",
            "tags": "微信客户",
            "type": "其他",
            "status": "待联系",
            "note": ("\n".join(lines))[:1000],
            "reminder_date": first_time[:10] if first_time else "",
        })
    return leads, None


def _row_to_lead(row, header_index):
    lead = {}
    for cell, idx in header_index.items():
        key = HEADER_MAP.get(str(cell).strip())
        if not key:
            continue
        val = row[idx] if idx < len(row) else None
        if val is None:
            val = ""
        lead[key] = str(val).strip()
    return lead


def parse_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], "文件为空"
    header = [str(c or "").strip() for c in rows[0]]
    header_index = {h: i for i, h in enumerate(header) if h in HEADER_MAP}
    if not header_index:
        return [], "第一行没有识别到表头（如：公司名称、联系人、电话）"
    leads = []
    for row in rows[1:]:
        lead = _row_to_lead(row, header_index)
        if lead.get("name"):
            leads.append(lead)
    return leads, None


def parse_csv(path):
    raw = open(path, "rb").read()
    text = None
    for enc in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return [], "无法识别文件编码"
    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader if any(str(c).strip() for c in r)]
    if not rows:
        return [], "文件为空"
    header = [str(c or "").strip() for c in rows[0]]
    header_index = {h: i for i, h in enumerate(header) if h in HEADER_MAP}
    if not header_index:
        return [], "第一行没有识别到表头（如：公司名称、联系人、电话）"
    leads = []
    for row in rows[1:]:
        lead = {}
        for h, idx in header_index.items():
            val = row[idx] if idx < len(row) else ""
            lead[HEADER_MAP[h]] = str(val).strip()
        if lead.get("name"):
            leads.append(lead)
    return leads, None


def parse_file(path, filename):
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".xlsx":
        return parse_xlsx(path)
    if ext in (".csv", ".txt"):
        return parse_csv(path)
    return [], "仅支持 .xlsx / .csv 文件"


def build_template_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "线索导入模板"
    ws.append(HEADERS)
    ws.append(["示例光纤科技有限公司", "张三", "13800138000", "zhang@example.com", "广东深圳", "工程商", "新线索", "批量导入", "光缆", "需要提供报价", "深圳市南山区"])
    ws.append(["", "", "", "", "", "", "", "", "", "", ""])
    widths = [28, 12, 15, 22, 12, 10, 10, 12, 12, 30, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _rows_to_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "客户线索"
    ws.append(HEADERS)
    for r in rows:
        ws.append([
            r.get("name", ""), r.get("contact", ""), r.get("phone", ""),
            r.get("email", ""), r.get("region", ""), r.get("type", ""),
            r.get("status", ""), r.get("source", ""), r.get("tags", ""),
            r.get("note", ""), r.get("address", ""),
        ])
    for i, w in enumerate([28, 12, 15, 22, 12, 10, 10, 12, 12, 30, 26], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_xlsx(rows):
    return _rows_to_workbook(rows).getvalue()


def export_csv(rows):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(HEADERS)
    for r in rows:
        writer.writerow([
            r.get("name", ""), r.get("contact", ""), r.get("phone", ""),
            r.get("email", ""), r.get("region", ""), r.get("type", ""),
            r.get("status", ""), r.get("source", ""), r.get("tags", ""),
            r.get("note", ""), r.get("address", ""),
        ])
    return "\ufeff" + buf.getvalue()
